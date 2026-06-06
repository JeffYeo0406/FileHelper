import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter


# Color palette — warm, elder-friendly
AMBER_FILL = PatternFill(start_color="F7E4BC", end_color="F7E4BC", fill_type="solid")
DARK_FILL = PatternFill(start_color="2C1A0E", end_color="2C1A0E", fill_type="solid")
GREEN_FILL = PatternFill(start_color="D4EDE1", end_color="D4EDE1", fill_type="solid")
ALT_ROW_FILL = PatternFill(start_color="FDFAF5", end_color="FDFAF5", fill_type="solid")
WHITE_FILL = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

DARK_FONT = Font(name="Calibri", size=14, color="2C1A0E")
WHITE_FONT = Font(name="Calibri", size=14, color="FFFFFF", bold=True)
AMBER_FONT = Font(name="Calibri", size=14, color="C8860A", bold=True)
HEADER_FONT = Font(name="Calibri", size=14, color="2C1A0E", bold=True)
TITLE_FONT = Font(name="Calibri", size=18, color="FFFFFF", bold=True)
METRIC_LABEL_FONT = Font(name="Calibri", size=11, color="6B5040")
METRIC_VALUE_FONT = Font(name="Calibri", size=14, color="2C1A0E", bold=True)

TABLE_BORDER = Border(
    left=Side(style="medium", color="A08070"),
    right=Side(style="medium", color="A08070"),
    top=Side(style="medium", color="A08070"),
    bottom=Side(style="medium", color="A08070")
)


def generate_two_tab_excel(
    rows: list[dict],
    template: dict,
    original_filename: str,
    output_path: str
) -> str:
    """Generate a 2-tab Excel workbook: Summary + Data."""
    os.makedirs(os.path.dirname(output_path), exist_ok=True)

    wb = Workbook()

    # ---- Summary Sheet ----
    ws_summary = wb.active
    ws_summary.title = "Summary"
    _build_summary(ws_summary, rows, template, original_filename)

    # ---- Data Sheet ----
    ws_data = wb.create_sheet("Data")
    _build_data(ws_data, rows, template)

    wb.save(output_path)
    return output_path


def _build_summary(ws, rows: list[dict], template: dict, filename: str):
    """Build the styled Summary sheet."""
    columns = template.get("columns", [])
    metrics = template.get("summary_metrics", [])

    # Header block
    ws.merge_cells("A1:C1")
    ws["A1"] = "Simple File Helper"
    ws["A1"].font = TITLE_FONT
    ws["A1"].fill = DARK_FILL
    for col in range(1, 4):
        ws.cell(row=1, column=col).fill = DARK_FILL

    # Metadata section
    row = 3
    meta = [
        ("Original File", filename),
        ("Template", template.get("name", "Unknown")),
        ("Processed On", datetime.now().strftime("%d %B %Y, %I:%M %p")),
        ("Records Found", str(len(rows))),
    ]
    for label, value in meta:
        ws.cell(row=row, column=1, value=label).font = METRIC_LABEL_FONT
        ws.cell(row=row, column=2, value=value).font = METRIC_VALUE_FONT
        ws.cell(row=row, column=1).fill = WHITE_FILL
        ws.cell(row=row, column=2).fill = WHITE_FILL
        row += 1

    # Metrics section
    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    ws.cell(row=row, column=1, value="Key Metrics").font = HEADER_FONT
    row += 1

    # Build column lookup for metrics
    col_lookup = {c["key"]: c["label"] for c in columns}

    for metric in metrics:
        col_key = metric["column"]
        agg = metric["agg"]
        label = metric["label"]

        if agg == "sum":
            value = sum(
                float(r.get(col_key, 0) or 0)
                for r in rows
            )
            formatted = f"MYR {value:,.2f}"
        elif agg == "count":
            value = len(rows)
            formatted = str(value)
        else:
            value = 0
            formatted = "—"

        ws.cell(row=row, column=1, value=label).font = METRIC_LABEL_FONT
        ws.cell(row=row, column=2, value=formatted).font = METRIC_VALUE_FONT
        ws.cell(row=row, column=1).fill = GREEN_FILL
        ws.cell(row=row, column=2).fill = GREEN_FILL
        ws.cell(row=row, column=1).border = TABLE_BORDER
        ws.cell(row=row, column=2).border = TABLE_BORDER
        row += 1

    # Column widths
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 36
    ws.column_dimensions["C"].width = 4


def _build_data(ws, rows: list[dict], template: dict):
    """Build the Data sheet with header row and all transaction rows."""
    columns = template.get("columns", [])

    # Header row
    ws.row_dimensions[1].height = 32
    for col_idx, col_def in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_def["label"])
        cell.font = Font(name="Calibri", size=14, color="2C1A0E", bold=True)
        cell.fill = AMBER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = TABLE_BORDER

    # Data rows
    for row_idx, record in enumerate(rows):
        excel_row = row_idx + 2
        ws.row_dimensions[excel_row].height = 36
        for col_idx, col_def in enumerate(columns, 1):
            key = col_def["key"]
            value = record.get(key, "")

            # Coerce types
            if col_def["type"] == "number" and value is not None:
                try:
                    value = float(value)
                except (ValueError, TypeError):
                    value = 0.0

            cell = ws.cell(row=excel_row, column=col_idx, value=value)

            # Alternating row shading
            if row_idx % 2 == 0:
                cell.fill = WHITE_FILL
            else:
                cell.fill = ALT_ROW_FILL

            cell.font = DARK_FONT
            cell.border = TABLE_BORDER
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

            # Number formatting for numeric columns
            if col_def["type"] == "number":
                cell.number_format = '#,##0.00'

    # Column widths
    for col_idx, col_def in enumerate(columns, 1):
        letter = get_column_letter(col_idx)
        ws.column_dimensions[letter].width = col_def.get("width", 16)

    # Freeze top row
    ws.freeze_panes = "A2"

    # Auto-filter
    last_col = get_column_letter(len(columns))
    ws.auto_filter.ref = f"A1:{last_col}{len(rows) + 1}"

    # Print settings — fit all columns on one page width, landscape
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_margins.left = 0.3
    ws.page_margins.right = 0.3
    ws.page_margins.top = 0.4
    ws.page_margins.bottom = 0.4
